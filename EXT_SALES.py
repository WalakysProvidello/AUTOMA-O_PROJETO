import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path

# =============================
# 1. CONFIGURAÇÃO DAS DATAS
# =============================
hoje = datetime.now()
ano = hoje.year
mes_num = hoje.strftime("%m")
dia_num = hoje.day

# Data para o nome do arquivo (ex: 2502)
data_nome_arquivo = hoje.strftime("%d%m")

meses_br = {
    "01": "JANEIRO", "02": "FEVEREIRO", "03": "MARÇO", "04": "ABRIL",
    "05": "MAIO", "06": "JUNHO", "07": "JULHO", "08": "AGOSTO",
    "09": "SETEMBRO", "10": "OUTUBRO", "11": "NOVEMBRO", "12": "DEZEMBRO"
}

# Padrões de pasta
nome_mes_extenso = meses_br[mes_num]
pasta_mes_leitura = f"{mes_num}.{nome_mes_extenso}"  # Ex: 02.FEVEREIRO
pasta_mes_saida = f"{mes_num} - {nome_mes_extenso}" # Ex: 02 - FEVEREIRO

# =============================
# 2. CAMINHOS PADRONIZADOS
# =============================
# Caminho de busca (Leitura)
base_dir = Path(fr'G:\MIS\MIS\Rafael Dias\Vend - Bases\SALES DEPOSITO\{ano}\{pasta_mes_leitura}')

# Caminho de saída (Rede G:)
dia_formatado = f"{dia_num:02d}"
saida_dir = Path(fr'G:\DIGITAL\COTA DISPARO MASSIVO\ALIMENTAÇÃO') / str(ano) / pasta_mes_saida / dia_formatado / "SALES"

# Cria a pasta de destino
saida_dir.mkdir(parents=True, exist_ok=True)

# Nome final: SALES_CONSOLIDADO_2502.xlsx
nome_arquivo_saida = f'SALES_CONSOLIDADO_{data_nome_arquivo}.xlsx'
caminho_final = saida_dir / nome_arquivo_saida

# =============================
# 3. FUNÇÕES DE APOIO
# =============================
def dentro_pasta_dia(root):
    """Verifica se a subpasta é de um dia anterior (D-1)"""
    try:
        caminho_rel = os.path.relpath(root, base_dir)
        partes = caminho_rel.split(os.sep)
        for parte in partes:
            if len(parte) == 5 and parte[2] == '.':
                dia_da_pasta = int(parte.split('.')[0])
                if dia_da_pasta < dia_num:
                    return True
    except:
        pass
    return False

def extrair_data(caminho_arquivo):
    """Extrai a data base do nome da pasta (ex: 24.02)"""
    partes = str(caminho_arquivo).split(os.sep)
    for parte in partes:
        if len(parte) == 5 and parte[2] == '.':
            try:
                d, m = parte.split('.')
                return f"{d}/{m}/{ano}"
            except:
                return None
    return None

# =============================
# 4. VARREDURA E CONSOLIDAÇÃO
# =============================
dados = []
extensoes_excel = ['.xlsx', '.xls', '.xlsm', '.csv']

print(f"Lendo arquivos de: {base_dir}")

for root, dirs, files in os.walk(base_dir):
    if dentro_pasta_dia(root):
        for file in files:
            if any(file.lower().endswith(ext) for ext in extensoes_excel) and not file.startswith('~$'):
                caminho_arquivo = Path(root) / file
                print(f"Processando: {file}")

                try:
                    if file.lower().endswith('.csv'):
                        df = pd.read_csv(caminho_arquivo, sep=None, engine='python')
                        linhas_com_conteudo = len(df)
                        valor_d2 = None
                    else:
                        wb = load_workbook(caminho_arquivo, data_only=True)
                        sheet = wb.active
                        
                        linhas_com_conteudo = 0
                        for row in sheet.iter_rows(values_only=True):
                            if any(cell not in (None, "") for cell in row):
                                linhas_com_conteudo += 1

                        valor_d2 = sheet['D2'].value
                        wb.close()

                    data_referencia = extrair_data(caminho_arquivo)
                    dados.append([str(caminho_arquivo), data_referencia, linhas_com_conteudo, valor_d2])

                except Exception as e:
                    print(f"Erro no arquivo {file}: {e}")

# =============================
# 5. SALVAMENTO
# =============================
if dados:
    df_final = pd.DataFrame(dados, columns=['Arquivo', 'Data', 'Qtde_Linhas', 'Valor_D2'])
    
    # Se já existir um consolidado do dia, removemos para evitar conflito
    if caminho_final.exists():
        caminho_final.unlink()
        
    df_final.to_excel(caminho_final, index=False)
    print(f"\n--- SUCESSO ---")
    print(f"Arquivo gerado: {nome_arquivo_saida}")
    print(f"Caminho: {saida_dir}")
else:
    print("\nAVISO: Nenhum arquivo D-1 foi encontrado para consolidar.")